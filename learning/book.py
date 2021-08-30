from openpyxl.workbook import Workbook
from openpyxl import load_workbook

wb = Workbook()
ws = wb.active

ws1 = wb.create_sheet('NewSheet')
ws2 = wb.create_sheet('Another', 0)

ws.title = 'MySheet'


wb2 = load_workbook('Exercise Files/regions.xlsx')

new_sheet = wb2.create_sheet('NewSheet')
active_sheet = wb2.active

cell = active_sheet['A1']

active_sheet['A1'] = 0

wb2.save('book.xlsx')
